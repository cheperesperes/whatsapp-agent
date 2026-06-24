#!/usr/bin/env python3
"""Generate the Oiikon wholesale catalog (.xlsx) with product images, LA-pickup price +5% markup.

Product names/specs come from the Oiikon project catalog (supabase/schema.sql, AGENT_PROMPT.md).
Product images are decoded from base64 pulled out of the Oiikon Supabase storage (saved to RESULT_FILE).
Source prices / MOQ / stock are the LA-warehouse pickup figures provided by the operator.
"""
import os, re, json, base64, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

RESULT_FILE = "/root/.claude/projects/-home-user-whatsapp-agent/159f3def-441c-553e-9106-56e043215e43/tool-results/mcp-Supabase-execute_sql-1782296545755.txt"
IMGDIR = "/tmp/claude-0/-home-user-whatsapp-agent/159f3def-441c-553e-9106-56e043215e43/scratchpad/img"
os.makedirs(IMGDIR, exist_ok=True)

# ---- decode images from the saved SQL result ----
raw = open(RESULT_FILE, encoding="utf-8").read()
try:
    raw = json.loads(raw)["result"]   # outer MCP envelope -> inner text
except (json.JSONDecodeError, KeyError, TypeError):
    pass
m = re.search(r"\[\s*\{.*\}\s*\]", raw, re.S)
rows = json.loads(m.group(0))
img_paths = {}
for row in rows:
    sku, b64 = row["sku"], row["b64"]
    img = PILImage.open(io.BytesIO(base64.b64decode(b64))).convert("RGB")
    img.thumbnail((150, 150), PILImage.LANCZOS)
    out = os.path.join(IMGDIR, f"{sku}.png")
    img.save(out, "PNG")
    img_paths[sku] = (out, img.size)
print("Decoded images:", ", ".join(sorted(img_paths)))

# Each item: (sku, full name, capacity, output, cost USD @ LA pickup, MOQ, stock)
DATA = [
    ("Portable Power Stations (LiFePO4)", [
        ("E2000LFP", "PECRON E2000LFP", "1,920 Wh", "2,000 W",            465, 24, 15),
        ("E3600LFP", "PECRON E3600LFP", "3,072 Wh", "3,600 W",            772, 25, 261),
        ("E3800LFP", "PECRON E3800LFP", "3,840 Wh", "4,200 W",            872, 25, 400),
        ("F3000LFP", "PECRON F3000LFP", "3,072 Wh", "3,600 W",            595, 30, 853),
        ("F5000LFP", "PECRON F5000LFP", "5,120 Wh", "7,200 W (120/240V)", 1100, 30, 139),
    ]),
    ("Expansion Batteries (LiFePO4)", [
        ("EP3800-48V", "PECRON EP3800-48V Expansion", "3,840 Wh", "48 V — pairs E3800/E3600/F3000/E2400", 490, 25, 600),
        ("FP5000-48V", "PECRON FP5000-48V Expansion", "5,120 Wh", "48 V — doubles F5000LFP autonomy",     850, 24, 438),
    ]),
    ("Portable Solar Panels", [
        ("PV200", "PECRON Flexible Monocrystalline 200W", "—", "200 W", 120, 40, 657),
        ("PV300", "PECRON Flexible Monocrystalline 300W", "—", "300 W", 180, 33, 171),
    ]),
]
MARKUP = 1.05  # +5% margin on LA-pickup cost

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Wholesale Catalog"

NAVY, BLUE, LIGHT, BAND, WHITE = "1F3864", "2E5496", "D9E1F2", "EFF3FA", "FFFFFF"
thin = Side(style="thin", color="B8C2D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Photo", "SKU", "Product", "Capacity", "Output", "List Price\n(LA Pickup)",
           "Wholesale\n(+5%)", "MOQ", "Min. Order\nValue", "Stock"]
NCOL = len(headers)
last_col = get_column_letter(NCOL)

# ---- Title block ----
ws.merge_cells(f"A1:{last_col}1")
c = ws["A1"]; c.value = "OIIKON — WHOLESALE CATALOG"
c.font = Font(size=22, bold=True, color=WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[1].height = 40

ws.merge_cells(f"A2:{last_col}2")
c = ws["A2"]; c.value = "Pick Up in LA Warehouse  •  All prices in USD  •  Wholesale price includes 5% margin"
c.font = Font(size=10, italic=True, color=WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor=BLUE)
ws.row_dimensions[2].height = 20

ws.merge_cells(f"A3:{last_col}3")
c = ws["A3"]; c.value = "Effective 2026-06-24  •  info@cbesoln.com"
c.font = Font(size=9, color="606060")
c.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[3].height = 16

# ---- Header row ----
hrow = 5
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=hrow, column=j, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[hrow].height = 30

money_cols = {6, 7, 9}
PHOTO_ROW_H = 84          # points
EMU_PER_PX = 9525
r = hrow + 1
for category, items in DATA:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    cc = ws.cell(row=r, column=1, value=category.upper())
    cc.font = Font(bold=True, color=NAVY, size=11)
    cc.fill = PatternFill("solid", fgColor=LIGHT)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cc.border = border
    ws.row_dimensions[r].height = 22
    r += 1

    for i, (sku, name, cap, out, cost, moq, stock) in enumerate(items):
        wholesale = round(cost * MARKUP, 2)
        moq_value = round(wholesale * moq, 2)
        band = WHITE if i % 2 == 0 else BAND
        # col 1 left blank for the photo
        row_vals = [None, sku, name, cap, out, cost, wholesale, moq, moq_value, stock]
        for j, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=band)
            if j == 2:
                cell.font = Font(bold=True, color=NAVY)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif j == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif j in money_cols:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if j == 7:
                    cell.font = Font(bold=True, color=BLUE)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = PHOTO_ROW_H

        if sku in img_paths:
            path, (w, h) = img_paths[sku]
            xim = XLImage(path)
            xim.width, xim.height = w, h          # keep native (already <=150px)
            xim.anchor = f"A{r}"                  # standard single-cell anchor
            ws.add_image(xim)
        r += 1

# ---- Footer note ----
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
cell = ws.cell(row=r, column=1,
               value="Notes:  Wholesale price = List Price (LA pickup) + 5% margin.  "
                     "MOQ = Minimum Order Quantity per model.  Min. Order Value = Wholesale Price × MOQ.  "
                     "Stock subject to prior sale — confirm availability at time of order.  "
                     "Product images & specs per Oiikon PECRON catalog.")
cell.font = Font(size=9, italic=True, color="606060")
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[r].height = 40

widths = [16, 13, 32, 11, 26, 13, 12, 7, 13, 9]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1

out = "docs/Oiikon_Wholesale_Catalog_2026-06-24.xlsx"
wb.save(out)
print("Saved", out)
